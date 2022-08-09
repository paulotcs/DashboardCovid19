#  https://betterprogramming.pub/3-simple-ways-to-download-files-with-python-569cb91acae6
#  https://stackoverflow.com/questions/53930645/extract-xlsx-workbook-file-metadata-properties-in-python-3-6
#  https://www.geeksforgeeks.org/reading-excel-file-using-python/
#  https://www.geeksforgeeks.org/python-program-to-read-excel-file-using-xlrd/

from openpyxl import load_workbook
wb = load_workbook('owid-covid-data.xlsx')
wb.properties
print(wb.properties)

import wget
import os
# Defina o arquivo remoto a ser baixado
remote_url = 'https://covid.ourworldindata.org/data/owid-covid-data.xlsx'
# Defina o nome do arquivo local para salvar dados
local_file = 'owid-covid-data.xlsx'

# Se já existir o arquivo owid-covid-data.xlsx, apague-o e baixe o novo
if local_file in os.listdir():
    os.remove(local_file)

# Baixa o arquivo
wget.download(remote_url, local_file)

# Verifique se o arquivo foi baixado
wb = load_workbook('owid-covid-data.xlsx')
wb.properties
print(wb.properties)
