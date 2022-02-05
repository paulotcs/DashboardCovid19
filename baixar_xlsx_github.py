#  Baixando o arquivo xlsx do github do Our World In Data, com dados sobre a Covid-19 no mundo.

# Importando as bibliotecas necessárias
import wget
import os

from openpyxl import load_workbook                                                # openpyxl é uma biblioteca para manipulação de arquivos xlsx
wb = load_workbook('owid-covid-data.xlsx')
remote_url = 'https://covid.ourworldindata.org/data/owid-covid-data.xlsx'         # Defina o arquivo remoto a ser baixado
local_file = 'owid-covid-data.xlsx'                                               # Defina o nome do arquivo local para salvar os dados

if local_file in os.listdir():                           # Se já existir o arquivo owid-covid-data.xlsx, apague-o e baixe o novo
    os.remove(local_file)

wget.download(remote_url, local_file)                    # Baixa o arquivo

wb = load_workbook('owid-covid-data.xlsx')               # wb armazena informações do arquivo baixado   
wb.properties
print("Arquivo baixado!! \n" , wb.properties)
